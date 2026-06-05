"""
api/v1/rapports.py — Routes rapports
GET  /rapports/                    → liste des rapports
POST /rapports/generer             → générer un rapport (PDF/CSV/Excel)
GET  /rapports/{id}/telecharger   → télécharger un rapport
"""
from datetime import datetime, timezone
from typing import Optional
from pathlib import Path
import csv
import io

from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from database.postgresql import get_db, AsyncSessionLocal
from models.rapport import Rapport, FormatRapport, TypeGeneration
from models.utilisateur import Utilisateur
from api.v1.utilisateurs import get_current_user

router = APIRouter()


# Schémas Pydantic
class RapportCreate(BaseModel):
    titre: str
    periode_debut: datetime
    periode_fin: datetime
    format: FormatRapport = FormatRapport.PDF


class RapportResponse(BaseModel):
    id: int
    titre: str
    periode_debut: datetime
    periode_fin: datetime
    format: FormatRapport
    type_generation: TypeGeneration
    date_generation: datetime
    chemin_fichier: Optional[str]


# GET /rapports/ — liste des rapports
@router.get(
    "/",
    response_model=list[RapportResponse],
    summary="Lister tous les rapports",
)
async def lister_rapports(
    db: AsyncSession = Depends(get_db),
    _: Utilisateur = Depends(get_current_user),
):
    """Retourne la liste de tous les rapports générés."""
    result = await db.execute(
        select(Rapport).order_by(Rapport.date_generation.desc())
    )
    rapports = result.scalars().all()
    return rapports


# POST /rapports/generer — générer un rapport
@router.post(
    "/generer",
    response_model=RapportResponse,
    summary="Générer un rapport",
)
async def generer_rapport(
    rapport_data: RapportCreate,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: Utilisateur = Depends(get_current_user),
):
    """
    Génère un rapport sur la période spécifiée.
    La génération effective se fait en arrière-plan.
    """
    # Normalise en naive UTC (strip tzinfo) pour correspondre aux colonnes
    # PostgreSQL TIMESTAMP WITHOUT TIME ZONE
    def _to_naive_utc(dt: datetime) -> datetime:
        if dt.tzinfo is not None:
            return dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt

    rapport = Rapport(
        titre=rapport_data.titre,
        periode_debut=_to_naive_utc(rapport_data.periode_debut),
        periode_fin=_to_naive_utc(rapport_data.periode_fin),
        format=rapport_data.format,
        type_generation=TypeGeneration.MANUEL,
        auteur_id=current_user.id,
    )
    
    db.add(rapport)
    await db.commit()
    await db.refresh(rapport)
    
    # Lancer la génération en arrière-plan avec sa PROPRE session DB
    background_tasks.add_task(_generer_fichier, rapport.id)
    
    return rapport


# GET /rapports/{id}/telecharger — télécharger un rapport
@router.get(
    "/{id}/telecharger",
    summary="Télécharger un rapport",
)
async def telecharger_rapport(
    id: int,
    db: AsyncSession = Depends(get_db),
    _: Utilisateur = Depends(get_current_user),
):
    """Retourne le fichier du rapport généré."""
    result = await db.execute(select(Rapport).where(Rapport.id == id))
    rapport = result.scalar_one_or_none()

    if not rapport:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")

    if not rapport.chemin_fichier:
        raise HTTPException(status_code=400, detail="Rapport pas encore généré")

    file_path = Path(rapport.chemin_fichier)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Fichier du rapport non trouvé")

    # Déterminer le type MIME selon le format
    media_type = "application/pdf" if rapport.format == FormatRapport.PDF else \
                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if rapport.format == FormatRapport.EXCEL else \
                 "text/csv"

    filename = f"rapport_{rapport.titre}_{rapport.date_generation.strftime('%Y%m%d_%H%M%S')}.{rapport.format.value.lower()}"

    return FileResponse(
        path=file_path,
        media_type=media_type,
        filename=filename
    )


async def _generer_fichier(rapport_id: int):
    """
    Génère le fichier du rapport en arrière-plan.
    Ouvre sa PROPRE session DB (la session de requête est déjà fermée).
    """
    import logging
    logger = logging.getLogger(__name__)

    async with AsyncSessionLocal() as db:
        result = await db.execute(select(Rapport).where(Rapport.id == rapport_id))
        rapport = result.scalar_one_or_none()

        if not rapport:
            logger.error(f"❌ Rapport {rapport_id} non trouvé")
            return

        try:
            Path("/app/rapports").mkdir(exist_ok=True)
            file_path = Path(f"/app/rapports/rapport_{rapport_id}.{rapport.format.value.lower()}")

            if rapport.format == FormatRapport.CSV:
                await _generer_csv(rapport, db, file_path)
            elif rapport.format == FormatRapport.EXCEL:
                await _generer_excel(rapport, db, file_path)
            elif rapport.format == FormatRapport.PDF:
                await _generer_pdf(rapport, db, file_path)

            rapport.chemin_fichier = str(file_path)
            await db.commit()
            logger.info(f"✅ Rapport {rapport_id} généré: {rapport.chemin_fichier}")

        except Exception as e:
            logger.error(f"❌ Erreur génération rapport {rapport_id}: {e}")
            rapport.chemin_fichier = None
            await db.commit()


async def _generer_csv(rapport: Rapport, db: AsyncSession, file_path: Path):
    """Génère un rapport au format CSV."""
    from models.alerte import Alerte
    from models.equipement import Equipement

    # Récupérer les données
    result_alertes = await db.execute(
        select(Alerte).where(
            Alerte.timestamp >= rapport.periode_debut,
            Alerte.timestamp <= rapport.periode_fin
        )
    )
    alertes = result_alertes.scalars().all()

    result_equipements = await db.execute(select(Equipement))
    equipements = result_equipements.scalars().all()

    with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=';')

        # En-tête
        writer.writerow(['sep=;'])
        writer.writerow(['RAPPORT', rapport.titre])
        writer.writerow(['Periode', rapport.periode_debut.strftime('%d/%m/%Y %H:%M'), 'a', rapport.periode_fin.strftime('%d/%m/%Y %H:%M')])
        writer.writerow(['Date generation', rapport.date_generation.strftime('%d/%m/%Y %H:%M:%S')])
        writer.writerow([])

        # Statistiques
        writer.writerow(['STATISTIQUES'])
        writer.writerow(["Nombre d'alertes", len(alertes)])
        writer.writerow(["Nombre d'equipements", len(equipements)])
        writer.writerow(['Alertes critiques', sum(1 for a in alertes if a.niveau.value == 'CRITIQUE')])
        writer.writerow(['Alertes WARNING', sum(1 for a in alertes if a.niveau.value == 'WARNING')])
        writer.writerow([])

        # Liste des alertes
        writer.writerow(['ID ALERTE', 'NIVEAU', 'EQUIPEMENT ID', 'CPU (%)', 'RAM (%)', 'DATE/HEURE', 'MESSAGE'])
        for alerte in alertes:
            writer.writerow([
                alerte.id,
                alerte.niveau.value,
                alerte.equipement_id or 'N/A',
                f'{alerte.valeur_cpu:.1f}' if alerte.valeur_cpu else '0.0',
                f'{alerte.valeur_ram:.1f}' if alerte.valeur_ram else '0.0',
                alerte.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
                alerte.message
            ])

        # Liste des equipements
        writer.writerow([])
        writer.writerow(['EQUIPEMENTS'])
        writer.writerow(['ID', 'HOSTNAME', 'IP', 'TYPE', 'STATUT', 'SEUIL CPU WARNING', 'SEUIL CPU CRITIQUE'])
        for eq in equipements:
            writer.writerow([
                eq.id,
                eq.hostname or 'N/A',
                eq.adresse_ip,
                eq.type.value if hasattr(eq.type, 'value') else str(eq.type),
                eq.statut.value if hasattr(eq.statut, 'value') else str(eq.statut),
                eq.seuil_cpu_warning,
                eq.seuil_cpu_critique,
            ])


async def _generer_excel(rapport: Rapport, db: AsyncSession, file_path: Path):
    """Génère un rapport au format Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from models.alerte import Alerte
    from models.equipement import Equipement

    from openpyxl.utils import get_column_letter

    DARK  = "0A1628"
    MID   = "1E3A5F"
    CYAN  = "00D4FF"
    WHITE = "FFFFFF"
    RED   = "CC2200"
    YELLOW= "F0A500"
    GREEN = "00AA55"
    LGREY = "F2F4F7"
    MGREY = "D0D7E2"

    wb = Workbook()

    # ── Feuille 1 : Résumé ──────────────────────────────────
    ws = wb.active
    ws.title = "Resume"

    result_alertes = await db.execute(
        select(Alerte).where(
            Alerte.timestamp >= rapport.periode_debut,
            Alerte.timestamp <= rapport.periode_fin
        )
    )
    alertes = result_alertes.scalars().all()
    result_equipements = await db.execute(select(Equipement))
    equipements = result_equipements.scalars().all()

    nb_critique = sum(1 for a in alertes if a.niveau.value == 'CRITIQUE')
    nb_warning  = sum(1 for a in alertes if a.niveau.value == 'WARNING')
    eq_en_ligne = sum(1 for e in equipements if e.statut.value == 'EN_LIGNE')

    def hdr(cell, txt, bg=MID, fg=WHITE, sz=11, bold=True):
        cell.value = txt
        cell.font  = Font(bold=bold, color=fg, size=sz)
        cell.fill  = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def val(cell, txt, bold=False, color="000000", bg=None):
        cell.value = txt
        cell.font  = Font(bold=bold, color=color)
        if bg:
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.alignment = Alignment(vertical="center")

    # Titre
    ws.merge_cells('A1:F1')
    hdr(ws['A1'], f"NETWATCH — {rapport.titre.upper()}", bg=DARK, sz=14)
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    hdr(ws['A2'], f"Periode: {rapport.periode_debut.strftime('%d/%m/%Y %H:%M')} → {rapport.periode_fin.strftime('%d/%m/%Y %H:%M')}", bg=MID, sz=10)

    ws.merge_cells('A3:F3')
    hdr(ws['A3'], f"Genere le: {rapport.date_generation.strftime('%d/%m/%Y a %H:%M:%S')} | NetWatch Platform v1.0 — FSBM Hassan II, Casablanca", bg=MID, sz=9)
    ws.row_dimensions[3].height = 18

    # KPI
    row = 5
    ws.merge_cells(f'A{row}:F{row}')
    hdr(ws[f'A{row}'], "RESUME EXECUTIF", bg=DARK, sz=11)
    ws.row_dimensions[row].height = 22
    row += 1

    kpis = [
        ("Equipements supervises", len(equipements), None),
        ("Equipements EN LIGNE",   eq_en_ligne,       GREEN if eq_en_ligne == len(equipements) else YELLOW),
        ("Total alertes",          len(alertes),       None),
        ("Alertes CRITIQUES",      nb_critique,        RED if nb_critique > 0 else GREEN),
        ("Alertes WARNING",        nb_warning,         YELLOW if nb_warning > 0 else GREEN),
    ]
    for label, valeur, color in kpis:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color=LGREY, end_color=LGREY, fill_type="solid")
        c = ws[f'B{row}']
        c.value = valeur
        c.font  = Font(bold=True, color=color or "000000")
        c.alignment = Alignment(horizontal="center")
        row += 1

    # ── Feuille 2 : Alertes ─────────────────────────────────
    ws2 = wb.create_sheet("Alertes")
    ws2.merge_cells('A1:G1')
    hdr(ws2['A1'], "ANALYSE DES ALERTES", bg=DARK, sz=12)
    ws2.row_dimensions[1].height = 25

    hdrs2 = ['ID', 'NIVEAU', 'EQUIPEMENT ID', 'CPU (%)', 'RAM (%)', 'DATE/HEURE', 'MESSAGE']
    for i, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=2, column=i)
        hdr(c, h, bg=MID)

    for row_i, alerte in enumerate(alertes, 3):
        bg = "FFE8E8" if alerte.niveau.value == 'CRITIQUE' else ("FFFDE8" if alerte.niveau.value == 'WARNING' else None)
        data = [
            alerte.id,
            alerte.niveau.value,
            alerte.equipement_id or 'N/A',
            round(alerte.valeur_cpu, 1) if alerte.valeur_cpu else 0.0,
            round(alerte.valeur_ram, 1) if alerte.valeur_ram else 0.0,
            alerte.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
            alerte.message,
        ]
        for col_i, v in enumerate(data, 1):
            c = ws2.cell(row=row_i, column=col_i, value=v)
            if bg:
                c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            c.alignment = Alignment(vertical="center")

    # ── Feuille 3 : Equipements ─────────────────────────────
    ws3 = wb.create_sheet("Equipements")
    ws3.merge_cells('A1:G1')
    hdr(ws3['A1'], "INVENTAIRE DES EQUIPEMENTS", bg=DARK, sz=12)
    ws3.row_dimensions[1].height = 25

    hdrs3 = ['ID', 'HOSTNAME', 'IP', 'TYPE', 'STATUT', 'SEUIL CPU WARN', 'SEUIL CPU CRIT']
    for i, h in enumerate(hdrs3, 1):
        c = ws3.cell(row=2, column=i)
        hdr(c, h, bg=MID)

    for row_i, eq in enumerate(equipements, 3):
        statut = eq.statut.value if hasattr(eq.statut, 'value') else str(eq.statut)
        bg = "E8FFE8" if statut == 'EN_LIGNE' else "FFE8E8"
        data = [
            eq.id,
            eq.hostname or 'N/A',
            eq.adresse_ip,
            eq.type.value if hasattr(eq.type, 'value') else str(eq.type),
            statut,
            eq.seuil_cpu_warning,
            eq.seuil_cpu_critique,
        ]
        for col_i, v in enumerate(data, 1):
            c = ws3.cell(row=row_i, column=col_i, value=v)
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            c.alignment = Alignment(vertical="center")

    # Largeurs colonnes
    for ws_sheet in [ws, ws2, ws3]:
        for col in ws_sheet.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws_sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    wb.save(file_path)


async def _generer_pdf(rapport: Rapport, db: AsyncSession, file_path: Path):
    """Génère un rapport au format PDF riche."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from models.alerte import Alerte, NiveauAlerte
    from models.equipement import Equipement

    doc = SimpleDocTemplate(
        str(file_path), pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm,
        title=rapport.titre, author="NetWatch Platform"
    )
    W = A4[0] - 4*cm
    styles = getSampleStyleSheet()
    
    # Couleurs du thème
    DARK_BLUE = colors.HexColor('#0a1628')
    MID_BLUE = colors.HexColor('#1e3a5f')
    CYAN = colors.HexColor('#00d4ff')
    WHITE = colors.white
    RED_ALERT = colors.HexColor('#cc2200')
    YELLOW_WARN = colors.HexColor('#f0a500')
    GREEN_OK = colors.HexColor('#00aa55')
    LIGHT_GREY = colors.HexColor('#f2f4f7')
    MED_GREY = colors.HexColor('#d0d7e2')

    s_title = ParagraphStyle('Title', fontSize=26, textColor=colors.HexColor('#00d4ff'), fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_sub = ParagraphStyle('Sub', fontSize=12, textColor=colors.HexColor('#00d4ff'), fontName='Helvetica', alignment=TA_CENTER, spaceAfter=6)
    s_h1 = ParagraphStyle('H1', fontSize=14, textColor=DARK_BLUE, fontName='Helvetica-Bold', spaceBefore=15, spaceAfter=10)
    s_body = ParagraphStyle('Body', fontSize=10, textColor=colors.black, fontName='Helvetica', spaceAfter=6, leading=14)

    def head_box(txt):
        t = Table([[Paragraph(txt, ParagraphStyle('HB', fontSize=12, textColor=WHITE, fontName='Helvetica-Bold'))]], colWidths=[W])
        t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), DARK_BLUE), ('PADDING', (0,0), (-1,-1), 8)]))
        return t

    result_alertes = await db.execute(
        select(Alerte).where(Alerte.timestamp >= rapport.periode_debut, Alerte.timestamp <= rapport.periode_fin).order_by(Alerte.timestamp.desc())
    )
    alertes = result_alertes.scalars().all()
    result_equipements = await db.execute(select(Equipement))
    equipements = result_equipements.scalars().all()

    nb_critique = sum(1 for a in alertes if a.niveau == NiveauAlerte.CRITIQUE)
    nb_warning = sum(1 for a in alertes if a.niveau == NiveauAlerte.WARNING)
    eq_en_ligne = sum(1 for e in equipements if e.statut.value == 'EN_LIGNE')

    story = []

    # --- COVER PAGE ---
    cover_bg = Table([[ 
        Paragraph("NETWATCH", s_title),
        Paragraph("Plateforme Intelligente de Supervision Réseau", s_sub),
        Spacer(1, 1*cm),
        Paragraph("RAPPORT DE SUPERVISION", ParagraphStyle('RT', fontSize=18, textColor=CYAN, alignment=TA_CENTER)),
        Paragraph(rapport.titre.upper(), ParagraphStyle('RN', fontSize=22, textColor=WHITE, alignment=TA_CENTER, spaceBefore=10)),
    ]], colWidths=[W])
    cover_bg.setStyle(TableStyle([
        ('BACKGROUND', (0,0),(-1,-1), DARK_BLUE),
        ('PADDING', (0,0),(-1,-1), 40),
        ('ROUNDEDCORNERS', [8]),
    ]))
    story.append(cover_bg)
    story.append(Spacer(1, 1*cm))

    meta_data = [
        ['Période analysée', f"{rapport.periode_debut.strftime('%d/%m/%Y %H:%M')}  →  {rapport.periode_fin.strftime('%d/%m/%Y %H:%M')}"],
        ['Date de génération', rapport.date_generation.strftime('%d/%m/%Y à %H:%M:%S')],
        ['Format', rapport.format.value],
    ]
    meta_table = Table(meta_data, colWidths=[5*cm, W-5*cm])
    meta_table.setStyle(TableStyle([
        ('FONTNAME', (0,0),(0,-1), 'Helvetica-Bold'),
        ('BACKGROUND',(0,0),(-1,-1), LIGHT_GREY),
        ('GRID', (0,0),(-1,-1), 0.5, MED_GREY),
        ('PADDING', (0,0),(-1,-1), 6),
    ]))
    story.append(meta_table)
    story.append(PageBreak())

    # --- 1. RESUME EXECUTIF ---
    story.append(head_box("1. RÉSUMÉ EXÉCUTIF"))
    story.append(Spacer(1, 0.5*cm))
    kpi_data = [
        ['INDICATEUR', 'VALEUR', 'ÉTAT'],
        ['Équipements supervisés', str(len(equipements)), '—'],
        ['Équipements EN LIGNE', str(eq_en_ligne), '✓ OK' if eq_en_ligne == len(equipements) else '⚠ Partiel'],
        ['Total alertes', str(len(alertes)), '—'],
        ['Alertes CRITIQUES', str(nb_critique), '🔴' if nb_critique > 0 else '✓'],
        ['Alertes WARNING', str(nb_warning), '🟡' if nb_warning > 0 else '✓'],
    ]
    kpi_table = Table(kpi_data, colWidths=[7*cm, 4*cm, W-11*cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0),(-1,0), MID_BLUE), ('TEXTCOLOR', (0,0),(-1,0), WHITE),
        ('FONTNAME', (0,0),(-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0),(-1,-1), 0.5, MED_GREY),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [WHITE, LIGHT_GREY]),
        ('PADDING', (0,0),(-1,-1), 6),
        ('ALIGN', (1,0),(-1,-1), 'CENTER'),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 1*cm))

    # --- 2. INVENTAIRE ---
    story.append(head_box("2. INVENTAIRE DES ÉQUIPEMENTS"))
    story.append(Spacer(1, 0.5*cm))
    eq_data = [['ID', 'IP', 'HOSTNAME', 'TYPE', 'STATUT']]
    for eq in equipements:
        eq_data.append([str(eq.id), eq.adresse_ip, eq.hostname or '—', eq.type.value if hasattr(eq.type, 'value') else str(eq.type), eq.statut.value if hasattr(eq.statut, 'value') else str(eq.statut)])
    eq_table = Table(eq_data, colWidths=[1.5*cm, 3.5*cm, 4.5*cm, 3*cm, 3.5*cm])
    eq_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0),(-1,0), MID_BLUE), ('TEXTCOLOR', (0,0),(-1,0), WHITE),
        ('FONTNAME', (0,0),(-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0),(-1,-1), 8),
        ('GRID', (0,0),(-1,-1), 0.5, MED_GREY), ('ROWBACKGROUNDS',(0,1),(-1,-1), [WHITE, LIGHT_GREY]),
        ('PADDING', (0,0),(-1,-1), 5), ('ALIGN', (0,0),(0,-1), 'CENTER'),
    ]))
    story.append(eq_table)
    story.append(Spacer(1, 1*cm))

    # --- 3. ALERTES ---
    story.append(head_box("3. ANALYSE DES ALERTES (50 Dernières)"))
    story.append(Spacer(1, 0.5*cm))
    if alertes:
        al_data = [['DATE/HEURE', 'NIVEAU', 'EQ_ID', 'CPU', 'RAM', 'SCORE IA']]
        for a in alertes[:50]:
            al_data.append([
                a.timestamp.strftime('%d/%m %H:%M'), a.niveau.value, str(a.equipement_id),
                f'{a.valeur_cpu:.1f}%', f'{a.valeur_ram:.1f}%', f'{a.score_anomalie:.3f}'
            ])
        al_table = Table(al_data, colWidths=[3*cm, 2.5*cm, 2*cm, 2*cm, 2*cm, 2.5*cm])
        al_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0),(-1,0), colors.HexColor('#2c3e50')), ('TEXTCOLOR', (0,0),(-1,0), WHITE),
            ('FONTNAME', (0,0),(-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0),(-1,-1), 8),
            ('GRID', (0,0),(-1,-1), 0.4, MED_GREY), ('ROWBACKGROUNDS',(0,1),(-1,-1), [WHITE, LIGHT_GREY]),
            ('PADDING', (0,0),(-1,-1), 5), ('ALIGN', (1,0),(-1,-1), 'CENTER'),
        ]))
        story.append(al_table)
    else:
        story.append(Paragraph("Aucune alerte enregistrée sur cette période.", s_body))
    
    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width=W, thickness=1, color=MID_BLUE))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph("NetWatch Platform v1.0 • PFE 2025/2026 — FSBM, Casablanca", ParagraphStyle('Footer', fontSize=7, textColor=colors.grey, alignment=TA_CENTER)))

    doc.build(story)


# GET /rapports/storage — taille du stockage /tmp
@router.get("/storage", summary="Taille stockage rapports")
async def storage_info(
    _: Utilisateur = Depends(get_current_user),
):
    tmp = Path("/app/rapports")
    tmp.mkdir(exist_ok=True)
    fichiers = list(tmp.glob("rapport_*"))
    taille_bytes = sum(f.stat().st_size for f in fichiers if f.exists())
    taille_mb = round(taille_bytes / (1024 * 1024), 2)
    return {
        "taille_mb": taille_mb,
        "nb_fichiers": len(fichiers),
        "max_mb": 1024,
        "pourcentage": round((taille_mb / 1024) * 100, 2)
    }


# GET /rapports/export-global — ZIP de tous les rapports
import zipfile
from fastapi.responses import StreamingResponse

@router.get("/export-global", summary="Exporter tous les rapports en ZIP")
async def export_global(
    _: Utilisateur = Depends(get_current_user),
):
    tmp = Path("/app/rapports")
    fichiers = list(tmp.glob("rapport_*"))
    fichiers = [f for f in fichiers if f.exists()]

    if not fichiers:
        raise HTTPException(status_code=404, detail="Aucun rapport disponible")

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in fichiers:
            zf.write(f, f.name)
    zip_buffer.seek(0)

    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=rapports_netwatch.zip"}
    )


# DELETE /rapports/{id} — supprimer un rapport
@router.delete("/{id}", summary="Supprimer un rapport")
async def supprimer_rapport(
    id: int,
    db: AsyncSession = Depends(get_db),
    _: Utilisateur = Depends(get_current_user),
):
    result = await db.execute(select(Rapport).where(Rapport.id == id))
    rapport = result.scalar_one_or_none()
    if not rapport:
        raise HTTPException(status_code=404, detail="Rapport non trouvé")
    # Supprimer le fichier physique si existe
    if rapport.chemin_fichier:
        f = Path(rapport.chemin_fichier)
        if f.exists():
            f.unlink()
    await db.delete(rapport)
    await db.commit()
    return {"message": f"Rapport #{id} supprimé"}
